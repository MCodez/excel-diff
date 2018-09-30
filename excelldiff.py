# -*- coding: utf-8 -*-
"""
Created on Sun Sep  9 23:51:34 2018

@author: LALIT ARORA
"""

import openpyxl
import datetime


def compare_workbook(a,b):
    wba=openpyxl.load_workbook(a)
    wbb=openpyxl.load_workbook(b)
    sheetsa=wba.sheetnames
    sheetsb=wbb.sheetnames
    uni_mismatch=0
    uni_mis=[]
    if (len(sheetsa)==len(sheetsb)) and (sheetsa==sheetsb):
        for i in range(len(sheetsa)):
            flags=compare_sheets(wba,wbb,sheetsa[i])
            uni_mismatch=uni_mismatch+flags[0]
            uni_mis.append(flags[1])
            if (flags[0]>0):
                print (str(flags[0])+" mismatches found in sheet : "+str(sheetsa[i]))
                print("Mismatch Data will be logged in Reports.")
            else:
                print("Sheet "+str(sheetsa[i])+" matched sucessfully.")
    else:
        uni_mismatch=1
    return (uni_mismatch,uni_mis)
        
def compare_sheets(a,b,sheet):
    ac_sheet_a=a[sheet]
    rowsa=ac_sheet_a.max_row
    colsa=ac_sheet_a.max_column
    
    ac_sheet_b=b[sheet]
    rowsb=ac_sheet_b.max_row
    colsb=ac_sheet_b.max_column
    
    rows=max(rowsa,rowsb)
    cols=max(colsa,colsb)
    
    mismatch=0
    mis=[]
    
    for i in range(1,rows+1):
        for j in range(1,cols+1):
            temp_a=str(ac_sheet_a.cell(row=i, column=j).value).strip()
            temp_b=str(ac_sheet_b.cell(row=i, column=j).value).strip()
            if (temp_a != temp_b):
                mismatch=mismatch+1
                m=str(temp_a)+" is replaced with "+str(temp_b)+" at coordinates ("+str(i)+" , "+str(j)+" )."
                mis.append(m)
                
    
    return (mismatch,mis)
            

def generate_report(wa,wb,mismatches,mis):
    if mismatches>0:
        now = datetime.datetime.now()
        time_stamp=str(now.year)+str(now.month)+str(now.hour)+str(now.minute)+str(now.second)+str(now.microsecond)
        rep_name = str(wa)+"_"+str(wb)+"_"+time_stamp+".txt"
        print("Mismatch report will be : "+rep_name)
        data=""
        for i in range(len(mis)):
            data=data+"\n\nSheet No. : "+str(i+1)
            data=data+"\n\n"
            data=data+"\n".join(mis[i])
            data=data+"\n----------------------------------------------------------"
            data=data+"\nTotal Mismatches in Sheet No : "+str(i+1)+"  :  "+str(len(mis[i]))
            data=data+"\n----------------------------------------------------------"
        data=data+"\n\n"
        data=data+"\n----------------------------------------------------------"
        data=data+"\n----------------------------------------------------------"
        data=data+"\nTotal Mismatches : "+str(mismatches)
        data=data+"\n----------------------------------------------------------"
        data=data+"\n----------------------------------------------------------"
        
        f=open(rep_name,"w")
        f.write(data)
        f.close()
            
if __name__=='__main__':
    import sys
    arguments=sys.argv
    result=compare_workbook(arguments[1],arguments[2])
    if result[0]==0:
        print ("Both Workbooks matched successfully.")
    else:
        if len(result[1])==0:
            print ("Workbook Sheet name error.")
        else:
            generate_report(arguments[1],arguments[2],result[0],result[1])



